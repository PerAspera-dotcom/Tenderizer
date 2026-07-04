"""Step 11 — end-to-end run / scheduler entrypoint.
Interface:
  run.run_pipeline(sources, db_path, out_path, tenant_id=DEFAULT_TENANT_ID,
                    now=None, fx_rates=None) -> health(dict)
  where each source = {"name":str, "fetch":callable->list[raw], "normalize":callable}
  A failing source must NOT abort the run; it is recorded in health.
  `fx_rates` defaults to a live ECB fetch (CR-001 D2) — tests pass a fixed
  snapshot so the suite doesn't depend on network access.
  `tenant_id` (Phase 2/3 step 3) scopes every store.py read/write.
"""
from openpyxl import load_workbook
import run
from conftest import TEST_TENANT_ID

FX_RATES = {"date": "2026-07-01", "rates": {"EUR": 1.0, "SEK": 11.23}}

def _src_ok(raw_ted_supply):
    return {"name":"TED","fetch":lambda:[raw_ted_supply],
            "normalize":__import__("normalize").normalize_ted}

def _src_boom():
    def boom(): raise RuntimeError("portal down")
    return {"name":"BROKEN","fetch":boom,"normalize":lambda r:r}

def test_end_to_end_produces_report(tmp_path, raw_ted_supply):
    health = run.run_pipeline([_src_ok(raw_ted_supply)],
                              str(tmp_path/"t.db"), str(tmp_path/"r.xlsx"),
                              tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)
    assert "ok" in health["TED"]
    assert load_workbook(str(tmp_path/"r.xlsx"))

def test_failing_source_is_isolated(tmp_path, raw_ted_supply):
    health = run.run_pipeline([_src_ok(raw_ted_supply), _src_boom()],
                              str(tmp_path/"t.db"), str(tmp_path/"r.xlsx"),
                              tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)
    assert "ok" in health["TED"]            # good source still ran
    assert "error" in health["BROKEN"]      # bad source captured, not fatal

def test_run_is_idempotent(tmp_path, raw_ted_supply):
    import store
    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    run.run_pipeline([_src_ok(raw_ted_supply)], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)
    run.run_pipeline([_src_ok(raw_ted_supply)], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)
    conn = store.init_db(db)
    assert len(store.all_records(conn, TEST_TENANT_ID)) == 1               # no duplicate

def test_excluded_notice_is_stored_but_not_reported(tmp_path, raw_ted_supply):
    # CR-001 F3: container/modular/prefab notices are auditable (kept in the DB with
    # exclude_reason set) but must not surface in the report.
    import store, copy
    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "999999-2026"
    raw["notice-title"] = {"eng": "Sweden – Supply of modular prefabricated cabins"}
    raw["classification-cpv"] = ["44211100"]
    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    conn = store.init_db(db)
    stored = store.all_records(conn, TEST_TENANT_ID)
    assert stored[0]["exclude_reason"] == "container_modular_prefab"  # auditable, not deleted

    wb = load_workbook(out)
    pub_numbers = [c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row]
    assert "999999-2026" not in pub_numbers                            # but not surfaced

def test_below_value_floor_is_excluded_after_fx_conversion(tmp_path, raw_ted_supply):
    # CR-001 F6: 100,000 SEK / 11.23 =~ 8,904 EUR — well under the 200k floor.
    import store, copy
    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "888888-2026"
    raw["estimated-value-proc"] = "100000"
    raw["estimated-value-cur-proc"] = "SEK"
    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    conn = store.init_db(db)
    stored = store.all_records(conn, TEST_TENANT_ID)[0]
    assert stored["exclude_reason"] == "below_value_floor"
    assert stored["fx_rate_date"] == "2026-07-01"   # conversion is reproducible (D2)

def test_above_value_floor_is_kept(tmp_path, raw_ted_supply):
    import store, copy
    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "777777-2026"
    raw["estimated-value-proc"] = "5000000"
    raw["estimated-value-cur-proc"] = "SEK"
    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    conn = store.init_db(db)
    assert store.all_records(conn, TEST_TENANT_ID)[0]["exclude_reason"] == ""

def test_republished_notices_collapse_via_full_pipeline(tmp_path, raw_ted_supply):
    # CR-001 D-DUP's own example: the same tender republished under a new
    # pub_number should collapse to one surfaced record.
    import store, copy
    raw_a = copy.deepcopy(raw_ted_supply)
    raw_a["publication-number"] = "111111-2026"
    raw_a["notice-title"] = {"eng": "Romania - Tents - Corturi si Generatoare"}
    raw_a["publication-date"] = "20260601"

    raw_b = copy.deepcopy(raw_ted_supply)
    raw_b["publication-number"] = "222222-2026"
    raw_b["notice-title"] = {"eng": "Romania - Tents - Corturi si Generatoare"}
    raw_b["publication-date"] = "20260615"   # republished two weeks later

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw_a, raw_b],
           "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    conn = store.init_db(db)
    records = {r["pub_number"]: r for r in store.all_records(conn, TEST_TENANT_ID)}
    assert records["111111-2026"]["exclude_reason"] == "superseded"
    assert records["222222-2026"]["supersedes"] == ["111111-2026"]

    wb = load_workbook(out)
    pub_numbers = [c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row]
    assert "111111-2026" not in pub_numbers      # collapsed, not shown twice
    assert "222222-2026" in pub_numbers          # the kept (latest) version surfaces

def test_genuinely_different_tenders_same_buyer_are_not_merged(tmp_path, raw_ted_supply):
    import store, copy
    raw_a = copy.deepcopy(raw_ted_supply)
    raw_a["publication-number"] = "333333-2026"
    raw_a["notice-title"] = {"eng": "Sweden - Supply of military tents"}

    raw_b = copy.deepcopy(raw_ted_supply)
    raw_b["publication-number"] = "444444-2026"
    raw_b["notice-title"] = {"eng": "Sweden - Catering services for field camps"}
    raw_b["classification-cpv"] = ["39522530", "55520000"]

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw_a, raw_b],
           "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    conn = store.init_db(db)
    records = {r["pub_number"]: r for r in store.all_records(conn, TEST_TENANT_ID)}
    assert records["333333-2026"]["exclude_reason"] == ""
    assert records["444444-2026"]["exclude_reason"] == ""


# ── CR-001 R3: translation (DeepL mocked — no real network call) ────────────

def test_non_english_notice_is_translated(tmp_path, raw_ted_supply, monkeypatch):
    import translate, copy
    monkeypatch.setattr(translate, "translate_to_english",
                         lambda text, api_key=None, timeout=15: (f"[EN] {text}", "ok"))
    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "555555-2026"
    del raw["notice-title"]["eng"]   # no English title -> normalize picks 'fra'
    raw["notice-title"]["fra"] = "Suède - Fourniture de tentes militaires"

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    import store
    stored = store.all_records(store.init_db(db), TEST_TENANT_ID)[0]
    assert stored["language"] == "fra"
    assert stored["translation_status"] == "ok"
    assert stored["tag_line_en"] == "[EN] Suède - Fourniture de tentes militaires"

def test_english_notice_is_never_sent_to_translate(tmp_path, raw_ted_supply, monkeypatch):
    import translate
    def boom(*a, **k):
        raise AssertionError("should not translate an already-English notice")
    monkeypatch.setattr(translate, "translate_to_english", boom)

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    run.run_pipeline([_src_ok(raw_ted_supply)], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    import store
    stored = store.all_records(store.init_db(db), TEST_TENANT_ID)[0]
    assert stored["language"] == "eng"
    assert stored["translation_status"] == ""

def test_excluded_notice_is_never_translated(tmp_path, raw_ted_supply, monkeypatch):
    # CR-001: translation runs after filters/dedup — don't spend DeepL quota
    # on a notice that won't surface anyway.
    import translate, copy
    def boom(*a, **k):
        raise AssertionError("should not translate an excluded notice")
    monkeypatch.setattr(translate, "translate_to_english", boom)

    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "666666-2026"
    del raw["notice-title"]["eng"]
    raw["notice-title"]["fra"] = "Location de tentes"   # F2 rental exclude

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    import store
    stored = store.all_records(store.init_db(db), TEST_TENANT_ID)[0]
    assert stored["exclude_reason"] == "rental"
    assert stored["translation_status"] == ""   # never attempted

def test_translation_failure_does_not_break_the_pipeline(tmp_path, raw_ted_supply, monkeypatch):
    import translate, copy
    monkeypatch.setattr(translate, "translate_to_english",
                         lambda text, api_key=None, timeout=15: (None, "unavailable"))
    raw = copy.deepcopy(raw_ted_supply)
    raw["publication-number"] = "999000-2026"
    del raw["notice-title"]["eng"]
    raw["notice-title"]["fra"] = "Suède - Fourniture de tentes militaires"

    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}
    health = run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)   # must not raise

    assert "ok" in health["TED"]
    import store
    stored = store.all_records(store.init_db(db), TEST_TENANT_ID)[0]
    assert stored["translation_status"] == "unavailable"
    assert stored["exclude_reason"] == ""   # a translation outage never excludes a tender
    assert load_workbook(out)               # report still builds fine


# ── Phase 2/3 step 3: tenant isolation through the full pipeline ────────────

def test_two_tenants_running_the_same_source_stay_isolated(tmp_path, raw_ted_supply):
    import store, copy
    raw = copy.deepcopy(raw_ted_supply)
    db = str(tmp_path/"t.db"); out1 = str(tmp_path/"r1.xlsx"); out2 = str(tmp_path/"r2.xlsx")
    src = {"name": "TED", "fetch": lambda: [raw], "normalize": __import__("normalize").normalize_ted}

    run.run_pipeline([src], db, out1, tenant_id=1, fx_rates=FX_RATES)
    run.run_pipeline([src], db, out2, tenant_id=2, fx_rates=FX_RATES)

    conn = store.init_db(db)
    assert len(store.all_records(conn, 1)) == 1
    assert len(store.all_records(conn, 2)) == 1


# ── Phase 2/3 step 5: per-tenant CPV/keywords/portals config ───────────────

def test_tenant_cpv_customization_affects_matching(tmp_path, raw_ted_supply):
    # raw_ted_supply's title mentions "tents" (a distinctive keyword) even
    # with an empty CPV set, so it still matches — just via keyword, not CPV.
    import store
    db = str(tmp_path/"t.db"); out = str(tmp_path/"r.xlsx")
    conn = store.init_db(db)
    store.ensure_tenant(conn, TEST_TENANT_ID)
    store.set_tenant_cpv(conn, TEST_TENANT_ID, [])
    src = {"name": "TED", "fetch": lambda: [raw_ted_supply], "normalize": __import__("normalize").normalize_ted}
    run.run_pipeline([src], db, out, tenant_id=TEST_TENANT_ID, fx_rates=FX_RATES)

    stored = store.all_records(conn, TEST_TENANT_ID)[0]
    assert stored["match_source"] == "keyword"


def test_default_sources_filters_by_enabled_portals(tmp_path):
    import store
    conn = store.init_db(str(tmp_path/"t.db"))
    store.ensure_tenant(conn, TEST_TENANT_ID)
    store.set_tenant_portal_enabled(conn, TEST_TENANT_ID, "BOAMP", False)

    sources = run._default_sources(conn, TEST_TENANT_ID, since=None)
    assert [s["name"] for s in sources] == ["TED"]
