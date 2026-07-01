"""Step 9 — report generator.
  report.build_report(records, health, out_path) -> out_path
  One sheet per section (Supply/Services/Works/Training/Other) + Health sheet.
  Records routed by category (Supply first); nothing filtered. Rows sorted by deadline.
"""
from openpyxl import load_workbook
import report

def _rec(pub, cat, deadline, tag="x"):
    return {"pub_number":pub,"tag_line":tag,"buyer":"b","country":"BE","category":cat,
            "procedure":"open","pub_date":"2026-06-01","deadline":deadline,
            "cpv_codes":["39522530"],"matched_terms":["tent"],"match_source":"both","url":"http://x"}

def test_report_file_created(tmp_path):
    out = report.build_report([_rec("1","Supply","2026-07-01")], {"TED":"ok"}, str(tmp_path/"r.xlsx"))
    assert load_workbook(out)

def test_sections_exist(tmp_path):
    wb = load_workbook(report.build_report([], {"TED":"ok"}, str(tmp_path/"r.xlsx")))
    for s in ["Supply","Services","Training"]:
        assert s in wb.sheetnames

def test_supply_record_lands_in_supply_sheet(tmp_path):
    wb = load_workbook(report.build_report([_rec("S1","Supply","2026-07-01")], {}, str(tmp_path/"r.xlsx")))
    assert "S1" in [c.value for row in wb["Supply"].iter_rows() for c in row]
    assert "S1" not in [c.value for row in wb["Services"].iter_rows() for c in row]

def test_services_record_lands_in_services_sheet(tmp_path):
    wb = load_workbook(report.build_report([_rec("SV1","Services","2026-07-01")], {}, str(tmp_path/"r.xlsx")))
    assert "SV1" in [c.value for row in wb["Services"].iter_rows() for c in row]

def test_works_record_lands_in_works_sheet(tmp_path):
    wb = load_workbook(report.build_report([_rec("W1","Works","2026-07-01")], {}, str(tmp_path/"r.xlsx")))
    assert "W1" in [c.value for row in wb["Works"].iter_rows() for c in row]

def test_unknown_category_lands_in_other(tmp_path):
    wb = load_workbook(report.build_report([_rec("O1","Banana","2026-07-01")], {}, str(tmp_path/"r.xlsx")))
    assert "O1" in [c.value for row in wb["Other"].iter_rows() for c in row]

def test_rows_sorted_by_deadline(tmp_path):
    recs = [_rec("late","Supply","2026-09-01"), _rec("early","Supply","2026-07-01")]
    wb = load_workbook(report.build_report(recs, {}, str(tmp_path/"r.xlsx")))
    assert [r[0].value for r in wb["Supply"].iter_rows(min_row=2)] == ["early","late"]

def test_empty_deadline_sorts_last(tmp_path):
    recs = [_rec("hasdate","Supply","2026-07-01"), _rec("nodate","Supply","")]
    wb = load_workbook(report.build_report(recs, {}, str(tmp_path/"r.xlsx")))
    assert [r[0].value for r in wb["Supply"].iter_rows(min_row=2)] == ["hasdate","nodate"]

def test_empty_input_is_valid_report(tmp_path):
    assert load_workbook(report.build_report([], {"TED":"ok"}, str(tmp_path/"r.xlsx")))

def test_health_sheet_records_status(tmp_path):
    wb = load_workbook(report.build_report([], {"TED":"ok (5)","BOAMP":"error: timeout"}, str(tmp_path/"r.xlsx")))
    vals = [c.value for row in wb["Health"].iter_rows() for c in row]
    assert "TED" in vals and "error: timeout" in vals
