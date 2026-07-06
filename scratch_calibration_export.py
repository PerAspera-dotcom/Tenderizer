"""One-off calibration export for a manual customer review session.

Two sheets:
  1. "Matched Tenders"   — everything currently surfaced (a real match_source,
     not excluded) — i.e. exactly what the customer sees today.
  2. "Excluded Sample"   — a stratified sample across the five exclusion
     reasons the customer asked to spot-check (rental, construction_works,
     container_modular_prefab, below_value_floor, deadline_too_soon) — NOT
     no_core_signal/superseded, which weren't in the requested scope. Each
     stratum is capped at 20 records (randomly sampled, fixed seed for a
     reproducible file) or taken in full if fewer exist, keeping the total
     in the ~50-100 range while still giving full visibility into the
     smaller categories rather than drowning them in construction_works.

Both sheets carry blank "Relevant (Y/N)" and "Notes" columns for the customer
to fill in. Read-only: reads the DB, writes an .xlsx, touches no filtering
logic or stored data.

Run from the project root:  python scratch_calibration_export.py [tenant_id]
"""
import random
import sys

sys.path.insert(0, "src")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import store
from report import _BORDER, _HEADER_FILL, _HEADER_FONT, _sort_key

OUT_PATH = "reports/calibration_review.xlsx"
SAMPLE_REASONS = ["construction_works", "deadline_too_soon", "below_value_floor",
                   "container_modular_prefab", "rental"]
PER_REASON_CAP = 20
SEED = 42

COLUMNS = [
    ("Pub. number", "pub_number"),
    ("Source", "source"),
    ("Category", "category"),
    ("Deadline", "deadline"),
    ("Tag line", None),          # resolved specially — translation-aware
    ("Language", "language"),
    ("Buyer", "buyer"),
    ("Country", "country"),
    ("CPV codes", "cpv_codes"),
    ("Matched terms", "matched_terms"),
    ("Match", "match_source"),
    ("Exclusion reason", "exclude_reason"),
    ("Link", "url"),
]
CALIBRATION_COLUMNS = ["Relevant (Y/N)", "Notes"]

_WIDTHS = {"Pub. number": 14, "Source": 9, "Category": 11, "Deadline": 20,
           "Tag line": 60, "Language": 10, "Buyer": 26, "Country": 9,
           "CPV codes": 20, "Matched terms": 22, "Match": 9,
           "Exclusion reason": 20, "Link": 38,
           "Relevant (Y/N)": 15, "Notes": 40}


def _tag_line(rec):
    # Same preference as the Review Queue UI: show the English translation
    # when it's ready, but keep the record's own language tagged alongside
    # (see "Language" column) rather than hiding the fact it was translated.
    if rec.get("translation_status") == "ok" and rec.get("tag_line_en"):
        return rec["tag_line_en"]
    return rec.get("tag_line") or ""


def _value(rec, label, key):
    if label == "Tag line":
        return _tag_line(rec)
    v = rec.get(key)
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return "" if v is None else v


def _write_sheet(ws, rows):
    all_columns = COLUMNS + [(c, None) for c in CALIBRATION_COLUMNS]
    for ci, (label, _) in enumerate(all_columns, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[c.column_letter].width = _WIDTHS.get(label, 18)
    for ri, rec in enumerate(sorted(rows, key=_sort_key), start=2):
        for ci, (label, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=ri, column=ci, value=_value(rec, label, key))
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(label == "Tag line"))
            if label == "Link" and rec.get("url"):
                cell.hyperlink = rec["url"]
                cell.font = Font(color="0563C1", underline="single")
        for j, _ in enumerate(CALIBRATION_COLUMNS, start=len(COLUMNS) + 1):
            ws.cell(row=ri, column=j).border = _BORDER
    ws.freeze_panes = "A2"


def stratified_excluded_sample(records, rng):
    by_reason = {r: [] for r in SAMPLE_REASONS}
    for rec in records:
        reason = rec.get("exclude_reason")
        if reason in by_reason:
            by_reason[reason].append(rec)

    sample = []
    print("Excluded-sample strata:")
    for reason in SAMPLE_REASONS:
        pool = by_reason[reason]
        k = min(PER_REASON_CAP, len(pool))
        chosen = rng.sample(pool, k) if pool else []
        sample.extend(chosen)
        print(f"  {reason:26s} available={len(pool):4d}  sampled={k}")
    return sample


def main():
    tenant_id = int(sys.argv[1]) if len(sys.argv) > 1 else 2
    conn = store.init_db("data/tenders.db")
    records = store.all_records(conn, tenant_id)

    matched = [r for r in records
               if r.get("match_source") in ("cpv", "both", "keyword")
               and not r.get("exclude_reason")]

    rng = random.Random(SEED)
    excluded_sample = stratified_excluded_sample(records, rng)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet("Matched Tenders"), matched)
    _write_sheet(wb.create_sheet("Excluded Sample"), excluded_sample)
    wb.save(OUT_PATH)

    print(f"\nMatched Tenders: {len(matched)} rows")
    print(f"Excluded Sample: {len(excluded_sample)} rows")
    print(f"Written to {OUT_PATH}")


if __name__ == "__main__":
    main()
