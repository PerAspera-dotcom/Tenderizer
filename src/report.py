"""Report generator — one sheet per category section + a Health sheet.

Divide, don't filter: every record is routed to its category's sheet (Supply first, as the
company's primary interest), with Works/Training/Other catching the rest so nothing is lost.
Rows are sorted by deadline (empty deadlines last). The Health sheet shows which sources ran.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Supply first (primary interest); Works/Training/Other ensure nothing is dropped.
SECTIONS = ["Supply", "Services", "Works", "Training", "Other"]

COLUMNS = [
    ("Pub. number", "pub_number"),
    ("Deadline", "deadline"),
    ("Tag line", "tag_line"),
    ("Buyer", "buyer"),
    ("Country", "country"),
    ("Procedure", "procedure"),
    ("CPV codes", "cpv_codes"),
    ("Matched terms", "matched_terms"),
    ("Match", "match_source"),
    ("Link", "url"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WIDTHS = {"Pub. number": 14, "Deadline": 20, "Tag line": 60, "Buyer": 28, "Country": 9,
           "Procedure": 14, "CPV codes": 22, "Matched terms": 26, "Match": 9, "Link": 40}


def _section_of(rec):
    cat = rec.get("category") or "Other"
    return cat if cat in SECTIONS else "Other"


def _value(rec, key):
    v = rec.get(key)
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return "" if v is None else v


def _sort_key(rec):
    d = rec.get("deadline") or ""
    return (d == "", d)   # empty deadlines sort last


def _write_sheet(ws, rows):
    for ci, (label, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[c.column_letter].width = _WIDTHS.get(label, 18)
    for ri, rec in enumerate(sorted(rows, key=_sort_key), start=2):
        for ci, (label, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=ri, column=ci, value=_value(rec, key))
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(label == "Tag line"))
            if label == "Link" and rec.get("url"):
                cell.hyperlink = rec["url"]
                cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"


def build_report(records, health, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    for sec in SECTIONS:
        ws = wb.create_sheet(sec)
        _write_sheet(ws, [r for r in records if _section_of(r) == sec])
    hs = wb.create_sheet("Health")
    for ci, label in enumerate(["Source", "Status"], start=1):
        c = hs.cell(row=1, column=ci, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    hs.column_dimensions["A"].width = 20
    hs.column_dimensions["B"].width = 40
    for ri, (src, status) in enumerate(health.items(), start=2):
        hs.cell(row=ri, column=1, value=src)
        hs.cell(row=ri, column=2, value=status)
    wb.save(out_path)
    return out_path
